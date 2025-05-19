from flask import Flask, render_template, request, redirect, url_for, send_file, session
import pandas as pd
import os
from datetime import datetime
import matplotlib.pyplot as plt
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'YourSuperSecretKey123'

DATA_FOLDER = "client_data"
STATIC_FOLDER = "static"
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(STATIC_FOLDER, exist_ok=True)

def age_group(age):
    if age is None:
        return "Unknown"
    elif age < 30:
        return "Under 30"
    elif age < 40:
        return "30-39"
    elif age < 50:
        return "40-49"
    elif age < 60:
        return "50-59"
    else:
        return "60+"

def recommend_goals(age_group, risk_profile):
    rules = {
        ('Under 30', 'High'): ["Wealth Creation", "Equity Investing", "International Travel"],
        ('30-39', 'Moderate'): ["Child Education", "Retirement Corpus", "Insurance Planning"],
        ('40-49', 'Low'): ["Debt Reduction", "Child Marriage", "Emergency Fund"],
        ('50-59', 'Moderate'): ["Retirement Planning", "Health Corpus", "Passive Income"],
        ('60+', 'Low'): ["Health Fund", "Fixed Income Stream", "Estate Planning"]
    }
    return rules.get((age_group, risk_profile), ["Basic Emergency Fund", "Insurance Review"])

def log_pdf_activity(client_file, action):
    with open("pdf_activity_log.csv", "a") as log:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        client_name = client_file.replace(".pdf", "").replace("client_", "").replace("_", " ").title()
        log.write(f"{client_name},{action},{timestamp}\n")

def generate_income_pie_chart(data):
    labels = []
    values = []

    income_fields = {
        "monthly_salary": "Monthly Salary",
        "business_income": "Business Income",
        "freelance_income": "Freelance",
        "rental_income": "Rental Income",
        "stock_dividend": "Stock Dividend",
        "online_courses": "Online Courses",
        "affiliate_marketing": "Affiliate Marketing",
        "digital_products": "Digital Products",
        "royalties": "Royalties",
        "app_sales": "Selling an App",
        "annuity_income": "Annuity Income",
        "gold_bond_income": "Sovereign Gold Bond",
        "reits_income": "REITs",
        "other_income": "Other"
    }

    # Try to get from data['income_sources'] if present, else from data directly
    income_data = data.get('income_sources', data)

    for key, label in income_fields.items():
        value = income_data.get(key, 0) or 0
        if value and value > 0:
            labels.append(label)
            values.append(value)

    if values:
        plt.figure(figsize=(4, 4))
        plt.pie(values, labels=labels, autopct="%1.1f%%", startangle=140)
        plt.title("Income Breakdown")
        chart_path = "static/income_pie_chart.png"
        plt.savefig(chart_path, bbox_inches="tight")
        plt.close()
        return chart_path
    return None

def generate_expense_pie_chart(data):
    labels = []
    values = []

    expense_fields = {
        "Housing": "Housing",
        "Food": "Food",
        "Transportation": "Transportation",
        "Utilities": "Utilities",
        "Entertainment": "Entertainment",
        "Loan EMI Payments": "Loan EMIs",
        "Life Insurance Premiums": "Life Insurance",
        "Health Insurance Premiums": "Health Insurance",
        "General Insurance Premiums": "General Insurance",
        "Mutual Funds": "Mutual Funds",
        "Recurring Deposits (RD)": "Recurring Deposits",
        "Clarity": "Clarity",
        "Others": "Others"
    }

    # Try to get from data['expense_breakdown'] if present, else from data directly
    expense_data = data.get('expense_breakdown', data)

    for key, label in expense_fields.items():
        value = expense_data.get(key, 0) or 0
        if value and value > 0:
            labels.append(label)
            values.append(value)

    if values:
        plt.figure(figsize=(4, 4))
        plt.pie(values, labels=labels, autopct="%1.1f%%", startangle=140)
        plt.title("Monthly Expenses Breakdown")
        chart_path = "static/expense_pie_chart.png"
        plt.savefig(chart_path, bbox_inches="tight")
        plt.close()
        return chart_path
    return None

def create_client_pdf(data, income_chart_path=None, expense_chart_path=None):
    pdf = FPDF()

    # Cover Page
    pdf.add_page()
    pdf.image("static/Round_Logo-removebg-preview.png", x=90, y=5, w=50)
    pdf.set_font("Arial", 'B', 18)
    pdf.set_text_color(0, 0, 120)
    pdf.ln(60)
    pdf.cell(200, 10, txt="Financial Health & Goal Discovery Report", ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(10)
    pdf.cell(200, 10, txt=f"Client Name: {data.get('full_name', '')}", ln=True, align='C')
    pdf.cell(200, 10, txt=f"Date: {datetime.now().strftime('%d %B %Y')}", ln=True, align='C')
    pdf.image("static/Signature.png", x=140, y=220, w=50)
    pdf.set_xy(140, 265)
    pdf.set_font("Arial", size=10)
    pdf.cell(50, 5, txt="Ranjit Chowdhury", ln=True, align='C')
    pdf.cell(50, 5, txt="Founder, Rudra Solutions", ln=True, align='C')

    # Summary Page
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.set_text_color(0, 0, 80)
    pdf.cell(200, 10, txt="Client Financial Summary", ln=True, align='C')
    pdf.ln(10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(200, 10, txt=f"Email: {data.get('email', '')}", ln=True)
    pdf.cell(200, 10, txt=f"Phone: {data.get('phone', '')}", ln=True)
    pdf.cell(200, 10, txt=f"City: {data.get('city', '')}", ln=True)
    pdf.cell(200, 10, txt=f"Age Group: {data.get('age_group', '')}", ln=True)
    pdf.cell(200, 10, txt=f"Risk Profile: {data.get('risk_profile', '')}", ln=True)
    pdf.ln(10)
    pdf.set_text_color(0, 0, 120)
    pdf.cell(200, 10, txt="Recommended Goals:", ln=True)
    pdf.set_text_color(0, 0, 0)
    for goal in data.get('recommended_goals', '').split(','):
        pdf.cell(200, 10, txt=f"- {goal.strip()}", ln=True)
    pdf.cell(200, 10, ln=True)
    pdf.cell(200, 10, txt="Income Details", ln=True, align="L")
    pdf.set_font("Arial", size=10)
    income_sources = data.get('income_sources', {})
    total_income = data.get('total_income', '')

    for source, amount in income_sources.items():
        pdf.cell(200, 8, txt=f"{source}: Rs.{amount:,.2f}", ln=True, align="L")

    pdf.set_font("Arial", "B", size=10)
    pdf.cell(200, 10, txt=f"Total Income: Rs.{total_income:,.2f}", ln=True, align="L")
    pdf.set_font("Arial", size=12)
    pdf.ln(5)
    if income_chart_path and os.path.exists(income_chart_path):
        pdf.image(income_chart_path, x=120, y=pdf.get_y() - 90, w=80)
    pdf.set_font("Arial", 'B', 12)
    pdf.ln(5)
    pdf.cell(200, 10, txt="Monthly Expenses", ln=True, align="L")
    pdf.set_font("Arial", size=10)

    expenses = data.get('expense_breakdown', {})
    for name, value in expenses.items():
        pdf.cell(200, 8, txt=f"{name}: Rs.{value:,.2f}", ln=True)

    pdf.set_font("Arial", 'B', size=10)
    pdf.cell(200, 10, txt=f"Total Expenses: Rs.{data.get('total_expenses', 0):,.2f}", ln=True)
    pdf.set_font("Arial", 'B', 12)
    pdf.ln(5)
    if expense_chart_path and os.path.exists(expense_chart_path):
        pdf.image(expense_chart_path, x=120, y=pdf.get_y() - 60, w=80)
    pdf.cell(200, 10, txt="Current Investments", ln=True, align="L")
    pdf.set_font("Arial", size=10)

    investments = data.get('investment_breakdown', {})
    for name, value in investments.items():
        pdf.cell(200, 8, txt=f"{name}: Rs.{value:,.2f}", ln=True)

    pdf.set_font("Arial", 'B', size=10)
    pdf.cell(200, 10, txt=f"Total Investments: Rs.{data.get('total_investments', 0):,.2f}", ln=True)
    pdf.cell(200, 10, txt=f"Emergency Fund Available?: {data.get('emergency_fund_status', 'No')}", ln=True)
    pdf.set_font("Arial", 'B', 12)
    pdf.ln(5)
    pdf.cell(200, 10, txt="Future Financial Goals", ln=True, align="L")
    pdf.set_font("Arial", size=10)

    pdf.cell(200, 8, txt=f"Retirement Goal: {data.get('retirement_goal', 'No')}", ln=True)
    if data.get("retirement_goal") == "Yes":
        pdf.cell(200, 8, txt=f"Retirement Corpus Amount: Rs.{data.get('retirement_amount', 0):,.2f}", ln=True)

    pdf.cell(200, 8, txt=f"Higher Education Planning: {data.get('education_goal', 'No')}", ln=True)
    if data.get("education_goal") == "Yes":
        pdf.cell(200, 8, txt=f"Education Fund Goal: Rs.{data.get('education_fund', 0):,.2f}", ln=True)

    pdf.multi_cell(200, 8, txt=f"Other Financial Goals: {data.get('other_goals', 'N/A')}")
    pdf.cell(200, 8, txt=f"Time Horizon for Primary Goal: {data.get('goal_horizon', 'N/A')}", ln=True)

    pdf.set_font("Arial", 'B', 12)
    pdf.ln(5)
    pdf.cell(200, 10, txt="Risk Capacity Assessment", ln=True, align="L")
    pdf.set_font("Arial", size=10)

    pdf.cell(200, 8, txt=f"Reaction to 20% Drop in Investments: {data.get('reaction_to_loss', '')}", ln=True)
    pdf.cell(200, 8, txt=f"Primary Financial Goal: {data.get('primary_goal', '')}", ln=True)
    pdf.cell(200, 8, txt=f"Income Stability: {data.get('income_stability', '')}", ln=True)
    pdf.cell(200, 8, txt=f"Has Liabilities: {data.get('has_liabilities', '')}", ln=True)
    pdf.cell(200, 8, txt=f"Monthly Income Saved: {data.get('monthly_savings', '')}", ln=True)

    pdf.set_font("Arial", 'B', 12)
    pdf.ln(5)
    pdf.cell(200, 10, txt="Consent", ln=True, align="L")
    pdf.set_font("Arial", size=10)

    pdf.cell(200, 8, txt=f"Consent to Share Financial Information: {data.get('consent_to_share', '')}", ln=True)

    filename = BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    filename.write(pdf_bytes)
    filename.seek(0)
    return filename

@app.route('/', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == 'admin' and password == 'securepass123':
            session['admin_logged_in'] = True
            return redirect(url_for('submit'))
        else:
            return "<h3 style='color:red;'>❌ Invalid credentials</h3>"
    return '''
        <h2>🔐 Admin Login</h2>
        <form method="POST">
            Username: <input type="text" name="username" required><br><br>
            Password: <input type="password" name="password" required><br><br>
            <input type="submit" value="Login">
        </form>
    '''

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/activity-log')
def activity_log():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    if not os.path.exists("pdf_activity_log.csv"):
        return "<p>No activity recorded yet.</p>"

    df = pd.read_csv("pdf_activity_log.csv", names=["Client", "Action", "Timestamp"])
    action_counts = df["Action"].value_counts()
    plt.figure(figsize=(4, 4))
    action_counts.plot.pie(autopct='%1.1f%%', startangle=90, title='PDF Engagement')
    chart_path = os.path.join(STATIC_FOLDER, "pdf_activity_chart_pie.png")
    plt.savefig(chart_path)
    plt.close()

    html_table = df.to_html(classes="table table-striped", index=False)

    return f"""
    <h2>📊 PDF Engagement Log</h2>
    <img src="/static/pdf_activity_chart_pie.png" width="300px">
    <br><br>
    {html_table}
    <br>
    <a href="/export-activity">
        <button>📥 Export to Excel</button>
    </a>
    """

@app.route('/export-activity')
def export_activity():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    log_path = "pdf_activity_log.csv"
    excel_path = "pdf_activity_report.xlsx"
    chart_image_path = os.path.join(STATIC_FOLDER, "pdf_activity_chart_bar.png")

    if not os.path.exists(log_path):
        return "<p>No activity log available to export.</p>"

    df = pd.read_csv(log_path, names=["Client", "Action", "Timestamp"])

    # Save Excel file
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Activity Log', index=False)

    # Create chart
    action_summary = df["Action"].value_counts()
    plt.figure(figsize=(6, 4))
    action_summary.plot(kind='bar', color='skyblue')
    plt.title("PDF Engagement Summary")
    plt.xlabel("Action")
    plt.ylabel("Count")
    plt.tight_layout()
    plt.savefig(chart_image_path)
    plt.close()

    # Insert chart into Excel
    wb = load_workbook(excel_path)
    ws = wb["Activity Log"]
    img = ExcelImage(chart_image_path)
    img.anchor = "E2"
    ws.add_image(img)
    wb.save(excel_path)

    return send_file(excel_path, as_attachment=True)

@app.route('/submit', methods=['GET', 'POST'])
def submit():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        city = request.form.get('city')
        age = request.form.get('age', type=int)
        risk_profile = request.form.get('risk_profile')

        age_grp = age_group(age)
        goals = recommend_goals(age_grp, risk_profile)

        income_sources = {
            "monthly_salary": request.form.get("monthly_salary", type=float),
            "business_income": request.form.get("business_income", type=float),
            "freelance_income": request.form.get("freelance_income", type=float),
            "rental_income": request.form.get("rental_income", type=float),
            "stock_dividend": request.form.get("stock_dividend", type=float),
            "online_courses": request.form.get("online_courses", type=float),
            "affiliate_marketing": request.form.get("affiliate_marketing", type=float),
            "digital_products": request.form.get("digital_products", type=float),
            "royalties": request.form.get("royalties", type=float),
            "app_sales": request.form.get("app_sales", type=float),
            "annuity_income": request.form.get("annuity_income", type=float),
            "gold_bond_income": request.form.get("gold_bond_income", type=float),
            "reits_income": request.form.get("reits_income", type=float),
            "other_income": request.form.get("other_income", type=float),
        }

        total_income = sum(income_sources.values())

        expense_data = {
            "Housing": request.form.get("housing_expense", type=float) or 0,
            "Food": request.form.get("food_expense", type=float) or 0,
            "Transportation": request.form.get("transportation_expense", type=float) or 0,
            "Utilities": request.form.get("utilities_expense", type=float) or 0,
            "Entertainment": request.form.get("entertainment_expense", type=float) or 0,
            "Loan EMI Payments": request.form.get("loan_emi_expense", type=float) or 0,
            "Life Insurance Premiums": request.form.get("life_insurance_expense", type=float) or 0,
            "Health Insurance Premiums": request.form.get("health_insurance_expense", type=float) or 0,
            "General Insurance Premiums": request.form.get("general_insurance_expense", type=float) or 0,
            "Mutual Funds": request.form.get("mutual_fund_expense", type=float) or 0,
            "Recurring Deposits (RD)": request.form.get("rd_expense", type=float) or 0,
            "Clarity": request.form.get("clarity_expense", type=float) or 0,
            "Others": request.form.get("other_expense", type=float) or 0,
        }

        total_expenses = sum(expense_data.values())

        investment_data = {
            "Total Equities": request.form.get("equities", type=float) or 0,
            "Total Debt Funds": request.form.get("debt_funds", type=float) or 0,
            "Real Estate": request.form.get("real_estate", type=float) or 0,
            "Certificate of Deposit": request.form.get("certificate_of_deposit", type=float) or 0,
            "Fixed Deposits (FD)": request.form.get("fd", type=float) or 0,
            "Recurring Deposits": request.form.get("rd", type=float) or 0,
            "Time Deposit (TD) Account": request.form.get("td", type=float) or 0,
            "National Savings Certificate (NSC)": request.form.get("nsc", type=float) or 0,
            "Public Provident Fund (PPF)": request.form.get("ppf", type=float) or 0,
            "Senior Citizens Savings Scheme": request.form.get("senior_savings", type=float) or 0,
            "National Pension System (NPS)": request.form.get("nps", type=float) or 0,
            "Mutual Funds": request.form.get("mutual_funds", type=float) or 0,
            "Silver ETFs": request.form.get("silver_etfs", type=float) or 0,
            "Gold ETFs": request.form.get("gold_etfs", type=float) or 0,
            "Sovereign Gold Bond": request.form.get("sovereign_gold_bond", type=float) or 0,
            "Life Insurance Policies (Total Sum Assured)": request.form.get("life_insurance_sum", type=float) or 0
        }

        emergency_fund_status = request.form.get("emergency_fund", "No")
        total_investments = sum(investment_data.values())

        client_data = {
            "full_name": full_name,
            "email": email,
            "phone": phone,
            "city": city,
            "age_group": age_grp,
            "risk_profile": risk_profile,
            "recommended_goals": ", ".join(goals), 
            "total_income": total_income
        }

        client_data['income_sources'] = income_sources
        client_data["expense_breakdown"] = expense_data
        client_data["total_expenses"] = total_expenses

        client_data["investment_breakdown"] = investment_data
        client_data["total_investments"] = total_investments
        client_data["emergency_fund_status"] = emergency_fund_status

        client_data["retirement_goal"] = request.form.get("retirement_goal", "No")
        client_data["retirement_amount"] = request.form.get("retirement_amount", type=float) or 0
        client_data["education_goal"] = request.form.get("education_goal", "No")
        client_data["education_fund"] = request.form.get("education_fund", type=float) or 0
        client_data["other_goals"] = request.form.get("other_goals", "")
        client_data["goal_horizon"] = request.form.get("goal_horizon", "")

        client_data["reaction_to_loss"] = request.form.get("reaction_to_loss", "")
        client_data["primary_goal"] = request.form.get("primary_goal", "")
        client_data["income_stability"] = request.form.get("income_stability", "")
        client_data["has_liabilities"] = request.form.get("has_liabilities", "")
        client_data["monthly_savings"] = request.form.get("monthly_savings", "")

        client_data["consent_to_share"] = request.form.get("consent_to_share", "")

        chart_path = generate_income_pie_chart(client_data)
        expense_chart_path = generate_expense_pie_chart(client_data)

        pdf_stream = create_client_pdf(client_data, income_chart_path=chart_path, expense_chart_path=expense_chart_path)
        log_pdf_activity(f"{full_name.replace(' ', '_')}_summary.pdf", "Generated")

        if chart_path and os.path.exists(chart_path):
            os.remove(chart_path)
        if expense_chart_path and os.path.exists(expense_chart_path):
            os.remove(expense_chart_path)

        return send_file(
            pdf_stream, 
            as_attachment=True, 
            download_name=f"{full_name.replace(' ', '_')}_summary.pdf", 
            mimetype="application/pdf"
        )

    return render_template("submit.html")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
